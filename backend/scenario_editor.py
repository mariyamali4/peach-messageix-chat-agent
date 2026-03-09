import pandas as pd
import numpy as np
import re
import os
from groq import Groq

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from backend.rag_core.retriever import retrieve_chunks


groq_api_key = os.environ.get("GROQ_API_KEY1")
client = Groq(api_key = groq_api_key)


def read_uploaded_file_data(file_path, query, embedding_model):
    xls = pd.ExcelFile(file_path)
    sheet_names = xls.sheet_names
    if len(sheet_names) > 1:
        q_emb = embedding_model.encode([query], convert_to_numpy=True, normalize_embeddings=True)
        sheet_embs = [embedding_model.encode([sheet_name], convert_to_numpy=True, normalize_embeddings=True) for sheet_name in sheet_names]

        scores = [np.dot(q_emb, sheet_emb.T).item() for sheet_emb in sheet_embs]
        best_idx = np.argmax(scores)
        best_sheet = sheet_names[best_idx]        
    else:
        best_sheet = sheet_names[0]
        
    print(f"Best matching sheet: {best_sheet}")
    df = xls.parse(best_sheet)
    return df, best_sheet

def run_scenario_agent(instruction, input_file, uploaded, output_file, embedding_model, index, metadata, max_retries=3):
    """
    Reads Excel, gets transformation code from model, executes it safely, saves new file.
    Returns structured output for front-end.
    Inputs:
    - instruction (str): User's instruction for Excel manipulation
    - uploaded (bool): Whether a file was uploaded in input
    - input_file (str): Path to input Excel file
    - output_file (str): Path to save updated Excel file
    - max_retries (int): Number of retries for code execution on failure

    Outputs:
    - dict with keys: 
        - success (bool), 
        - code (str), 
        - logs (str), 
        - downloadable modified excel file saved to output_file path
    """
    
    logs = []

    df_input, target_sheet_name = None, None
  #  if input_file is not None:
    if uploaded:
        logs.append(f"Using uploaded file: {input_file}")
        df_input, target_sheet_name = read_uploaded_file_data(input_file, instruction, embedding_model)
        logs.append(f"🔍 Identified target sheet: '{target_sheet_name}'")
    else:
        retriever_query = f"which MESSAGEix-Pakistan-CurPol sheet has information about this query: {instruction}"
        results = retrieve_chunks(retriever_query, embedding_model, index, metadata, k=1, for_rag=True)
     #   print(results['body'][0])
        target_sheet_name = results['body'][0].split('\n')[0]
        target_sheet_name = target_sheet_name.replace('Sheet: ', '')
        logs.append(f"🔍 Identified target sheet: '{target_sheet_name}'")

        xls = pd.ExcelFile(input_file)
        for sheet_name in xls.sheet_names:
            if sheet_name == target_sheet_name:
                df_input = xls.parse(sheet_name)
                break
        #logs.append(f"🔍 Identified target sheet: '{target_sheet_name}'")


    if df_input is None:
        raise ValueError(f"❌ No sheet named '{target_sheet_name}' found in {input_file}.")

    logs.append("📄 Loaded Excel file successfully.")
    logs.append(f"Columns: {list(df_input.columns)}")

    # Prepare prompt
    prompt = f"""
        You are a data engineer working with climate scenario data.
        You are given a pandas DataFrame named `df`.

        Schema:
        {list(df_input.columns)}

        Sample rows:
        {df_input.head().to_dict(orient="records")}

        Instruction:
        {instruction}

        TASK:
        Write Python (pandas and numpy) code that applies the instruction by modifying `df` in-place.

        LOGIC RULES (strict):
        1. Apply any temporal filters (e.g. "after 2030") BEFORE analysis.
        2. For "most/least expensive", compute the highest/lowest MEAN value unless stated otherwise.
        3. Identify technologies or categories by name/ID — NEVER by float value matching.
        4. Scope → aggregate → modify (in that order).
        5. Use vectorized operations only (no loops, no `.apply`).

        
        CODING RULES:
        - Modify `df` in-place using `.loc[...]`.
        - Use `.str.contains(..., case=False, na=False)`, , not exact matches, for string filters.
        - Preserve all rows/columns unless explicitly instructed to drop them.
        - Sort by time columns (e.g. `year`, `year_vtg`) if trends are implied.
        - Drop rows only via boolean indexing or `df.drop(...)`.    

        
        FORBIDDEN:
        - File I/O, system calls, env access.
        - Defining functions/classes.
        - Using os, sys, pathlib, subprocess, eval, exec.
        - Any code that triggers `SettingWithCopyWarning`.

        OUTPUT:
        - Return ONLY valid Python code
        - No explanations
        - No markdown
    """



    def generate_code(extra_context=None):
        context = prompt
        if extra_context:
            context += f"\nFix the issue described here: {extra_context}"

        completion = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {
                    "role": "user",
                    "content": prompt
                }
            ]
        )
        response = completion.choices[0].message.content
        return re.sub(r"^```(?:python)?|```$", "", response.strip(), flags=re.MULTILINE).strip()

    # First attempt
    code = generate_code()
    logs.append("🧠 Model-generated code:")
    logs.append(code)

    # Safety checks
    forbidden_patterns = [
        r"os\.", r"sys\.", r"open\s*\(", r"subprocess",
        r"eval\s*\(", r"exec\s*\(", r"__", r"shutil", r"pathlib"
    ]
    if any(re.search(p, code, re.IGNORECASE) for p in forbidden_patterns):
        print(code)
        raise ValueError("⚠️ Unsafe code detected! Execution blocked.")
    
    # Whiteliseted imports
    #allowed_imports = ["import numpy", "import numpy as np", "import pandas", "import pandas as pd"]

    import_lines = re.findall(r"^\s*import\s+[^\n]+", code, flags=re.MULTILINE)
    for line in import_lines:
        if not any(pkg in line for pkg in ["numpy", "pandas"]):
            raise ValueError(f"⚠️ Unsafe import detected: '{line.strip()}' — only numpy and pandas are allowed.")

    # --- Auto-inject safe imports if missing ---
    if "import pandas" not in code:
        logs.append("ℹ️ Auto-added: import pandas as pd")
        code = "import pandas as pd\n" + code
    if "import numpy" not in code:
        logs.append("ℹ️ Auto-added: import numpy as np")
        code = "import numpy as np\n" + code

    # Try executing
    for attempt in range(max_retries + 1):
        try:
            local_env = {"df": df_input.copy(), "pd": pd}
            exec(code, {}, local_env)
            df_new = local_env.get("df")

            if not isinstance(df_new, pd.DataFrame):
                raise ValueError("No valid DataFrame 'df' produced.")

            # --- Saving the edited sheet to the output Excel file, overwriting the target sheet ---
            # 1. Loading the existing workbook
            template_wb = load_workbook(input_file)
            
            # 2. Get the specific sheet (this preserves its position in the tab order)
            if target_sheet_name in template_wb.sheetnames:
                ws = template_wb[target_sheet_name]
                
                # 3. Clear the existing content (delete_rows from 1 to max_row ensures we don't leave old data behind)
                ws.delete_rows(1, ws.max_row)
            else:
                # Edge case: If it's a brand new sheet, it will go to the end
                ws = template_wb.create_sheet(title=target_sheet_name)

            # 4. Write the new data into the existing sheet
            for r in dataframe_to_rows(df_new, index=False, header=True):
                ws.append(r)

            # 5. Save the edited workbook to the output file path
            template_wb.save(output_file)

            logs.append(f"✅ Overwrote '{target_sheet_name}' and saved to {output_file}")

            return {"success": True, "code": code, "logs": "\n".join(logs)}

        except Exception as e:
            logs.append(f"❌ Error executing code: {e}")
            if attempt < max_retries:
                logs.append("🔁 Retrying with fix...")
                code = generate_code(extra_context=str(e))
            else:
                return {"success": False, "code": code, "logs": "\n".join(logs)}
